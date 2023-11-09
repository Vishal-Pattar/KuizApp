import openpyxl
import mysql.connector
import string
import random

############################################################################################
############################################################################################
db_table_name = "set_ai"
file_name = "questions.xlsx"
Sheet_name = "Sheet2"
############################################################################################
############################################################################################

mydb = mysql.connector.connect(host="localhost", user="root", password="", database="kuiz")
mycursor = mydb.cursor()

sqld = """CREATE TABLE IF NOT EXISTS %s (q_id INT AUTO_INCREMENT PRIMARY KEY, q_unique VARCHAR(8) NOT NULL,
 quno VARCHAR(10) NOT NULL, ques VARCHAR(100) NOT NULL, ans VARCHAR(5) NOT NULL, 
opt1 VARCHAR(50) NOT NULL, opt2 VARCHAR(50) NOT NULL, opt3 VARCHAR(50) NOT NULL, opt4 VARCHAR(50) NOT NULL)"""
#table = "q_set_prac"
table = db_table_name
mycursor.execute(sqld %table)
mydb.commit()
print("Database created")

path = file_name
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj[Sheet_name]

xx = 1
while True:
    cell_obj = sheet_obj.cell(row = xx, column = 2)
    aa = cell_obj.value
    if aa == None:
        break
    else:
        pass
    xx += 4

for x in range(1,xx,4):
    res = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
    sql1 = "INSERT INTO "+ table +" (q_unique) VALUES (%s)"
    val1 = (res,)
    mycursor.execute(sql1, val1)
    mydb.commit()
    if(x % 4 == 1):
        for y in range(1,3):
            cell_obj = sheet_obj.cell(row = x, column = y)
            cell = cell_obj.value
            if y == 1:
                sql2 = "UPDATE "+ table +" SET quno = %s WHERE q_unique = %s"
                val2 = (cell, res)
                mycursor.execute(sql2, val2)
                mydb.commit()
            elif y == 2:
                sql3 = "UPDATE "+ table +" SET ques = %s WHERE q_unique = %s"
                val3 = (cell, res)
                mycursor.execute(sql3, val3)
                mydb.commit()
    if((x+2) % 4 == 3):
        for z in range(1,6):
            cell_obj = sheet_obj.cell(row = (x+2), column = z)
            cell = cell_obj.value
            if z == 1:
                sql4 = "UPDATE "+ table +" SET ans = %s WHERE q_unique = %s"
                val4 = (cell, res)
                mycursor.execute(sql4, val4)
                mydb.commit()
                # print("val4",val4)
            elif z == 2:
                sql5 = "UPDATE "+ table +" SET opt1 = %s WHERE q_unique = %s"
                val5 = (cell, res)
                mycursor.execute(sql5, val5)
                mydb.commit()
                # print("val5",val5)
            elif z == 3:
                sql6 = "UPDATE "+ table +" SET opt2 = %s WHERE q_unique = %s"
                val6 = (cell, res)
                mycursor.execute(sql6, val6)
                mydb.commit()
                # print("val6",val6)
            elif z == 4:
                sql7 = "UPDATE "+ table +" SET opt3 = %s WHERE q_unique = %s"
                val7 = (cell, res)
                mycursor.execute(sql7, val7)
                mydb.commit()
                # print("val7",val7)
            elif z == 5:
                sql8 = "UPDATE "+ table +" SET opt4 = %s WHERE q_unique = %s"
                val8 = (cell, res)
                mycursor.execute(sql8, val8)
                mydb.commit()
                # print("val8",val8)
mydb.close()
print(int((xx-1)/4),"Records inserted Successfully")