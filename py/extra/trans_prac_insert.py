import openpyxl
import mysql.connector
import string
import random

mydb = mysql.connector.connect(host="localhost", user="root", password="", database="kuiz")
mycursor = mydb.cursor()

sqld = """CREATE TABLE IF NOT EXISTS %s (q_id INT AUTO_INCREMENT PRIMARY KEY, q_unique VARCHAR(8) NOT NULL,
 quno VARCHAR(10) NOT NULL, ques VARCHAR(100) NOT NULL, ans VARCHAR(5) NOT NULL, 
opt1 VARCHAR(50) NOT NULL, opt2 VARCHAR(50) NOT NULL, opt3 VARCHAR(50) NOT NULL, opt4 VARCHAR(50) NOT NULL)"""
table = "q_set_prac"
mycursor.execute(sqld %table)
mydb.commit()
print("Database created")

path = "questions.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj["Sheet2"]

xx = 1
while True:
    cell_obj = sheet_obj.cell(row = xx, column = 2)
    aa = cell_obj.value
    if aa == None:
        break
    else:
        pass
    xx += 4

for x in range(1,xx,4):
    res = (x+3)/4
    if(x % 4 == 1):
        for y in range(1,3):
            cell_obj = sheet_obj.cell(row = x, column = y)
            cell = cell_obj.value
            if y == 1:
                sql2 = "INSERT INTO "+ table +"(quno) VALUES (%s)"
                val2 = (cell,)
                mycursor.execute(sql2, val2)
                mydb.commit()
            elif y == 2:
                sql3 = "INSERT INTO "+ table +"(ques) VALUES (%s)"
                val3 = (cell,)
                mycursor.execute(sql3, val3)
                mydb.commit()
    if((x+2) % 4 == 3):
        for z in range(1,9):
            cell_obj = sheet_obj.cell(row = (x+2), column = z)
            cell = cell_obj.value
            if z == 1:
                sql4 = "INSERT INTO "+ table +"(ans) VALUES (%s)"
                val4 = (cell,)
                mycursor.execute(sql4, val4)
                mydb.commit()
                # print("val4",val4)
            elif z == 2:
                sql5 = "INSERT INTO "+ table +"(opt1) VALUES (%s)"
                val5 = (cell,)
                mycursor.execute(sql5, val5)
                mydb.commit()
                # print("val5",val5)
            elif z == 4:
                sql6 = "INSERT INTO "+ table +"(opt2) VALUES (%s)"
                val6 = (cell,)
                mycursor.execute(sql6, val6)
                mydb.commit()
                # print("val6",val6)
            elif z == 6:
                sql7 = "INSERT INTO "+ table +"(opt3) VALUES (%s)"
                val7 = (cell,)
                mycursor.execute(sql7, val7)
                mydb.commit()
                # print("val7",val7)
            elif z == 8:
                sql8 = "INSERT INTO "+ table +"(opt4) VALUES (%s)"
                val8 = (cell,)
                mycursor.execute(sql8, val8)
                mydb.commit()
                # print("val8",val8)
mydb.close()
print(int((xx-1)/4),"Records inserted Successfully")