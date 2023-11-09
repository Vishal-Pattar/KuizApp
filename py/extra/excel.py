import openpyxl

path = "questions.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj["Sheet1"]

for x in range(1,41,4):
    lst = []
    if(x % 4 == 1):
        for y in range(1,3):
            cell_obj = sheet_obj.cell(row = x, column = y)
            cell = cell_obj.value
            lst.append(cell)
    if((x+2) % 4 == 3):
        for z in range(1,6):
            cell_obj = sheet_obj.cell(row = (x+2), column = z)
            cell = cell_obj.value
            lst.append(cell)
    print(lst)

#print(cell_obj.value)


