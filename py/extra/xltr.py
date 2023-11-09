import openpyxl

path = "questions.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj["Sheet2"]

for x in range(1,41,4):
    res = int((x+3)/4)
    if(x % 4 == 1):
        qq = "Question no. " + str(res)
        cell_obj = sheet_obj.cell(row = x, column = 2)
        cell_obj.value = qq
        wb_obj.save(path)
    if((x+2) % 4 == 3):
        for z in range(2,9,2):
            aa = "opt" + str(int(z/2))
            cell_obj = sheet_obj.cell(row = (x+2), column = z)
            cell_obj.value = aa
            wb_obj.save(path)
            